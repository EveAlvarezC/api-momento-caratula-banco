import streamlit as st
import json, io, pathlib, zipfile
import pandas as pd
import fitz
from PIL import Image
from google import genai
from google.genai import types
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

# ── Configuración de página ──────────────────────────────────────────
st.set_page_config(page_title="Extractor de Cuentas Bancarias", layout="wide")
st.title("📄 Extractor de Datos Bancarios")
st.write("Sube los PDFs de carátulas bancarias y descarga un Excel con los datos extraídos.")

# ── API Key desde secrets ────────────────────────────────────────────
if "GEMINI_API_KEY" not in st.secrets:
    st.error("⚠️ Falta configurar la API Key de Gemini en los Secrets de la app.")
    st.stop()

client = genai.Client(api_key=st.secrets["GEMINI_API_KEY"])

# ── Prompt para Gemini ───────────────────────────────────────────────
PROMPT = """
Analiza este documento bancario y extrae la siguiente información.
Para cada campo devuelve el valor y el bounding box donde aparece en la imagen
(formato [ymin, xmin, ymax, xmax] normalizados de 0 a 1000).
Si un dato no aparece usa "Verificar" como valor y [0,0,0,0] como bbox.
Devuelve ÚNICAMENTE este JSON sin texto adicional ni bloques de código:
{
  "nombre_completo": {"valor": "...", "bbox": [ymin, xmin, ymax, xmax]},
  "cuenta":          {"valor": "...", "bbox": [ymin, xmin, ymax, xmax]},
  "clabe":           {"valor": "...", "bbox": [ymin, xmin, ymax, xmax]},
  "banco":           {"valor": "...", "bbox": [ymin, xmin, ymax, xmax]},
  "tipo":            {"valor": "TC o TD", "bbox": [ymin, xmin, ymax, xmax]}
}
tipo: TC = Tarjeta de Crédito, TD = Tarjeta de Débito / Cuenta de Débito.
"""


# ── Funciones auxiliares ─────────────────────────────────────────────
def pdf_primera_pagina(pdf_bytes):
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    pix = doc[0].get_pixmap(matrix=fitz.Matrix(3, 3))
    img_bytes = pix.tobytes("png")
    doc.close()
    return img_bytes


def recortar_zona(img_bytes, bbox, margen_h=150, margen_v=100, min_w=600, min_h=150):
    ymin, xmin, ymax, xmax = bbox
    if [ymin, xmin, ymax, xmax] == [0, 0, 0, 0]:
        return None

    img = Image.open(io.BytesIO(img_bytes))
    w, h = img.size

    x0 = int(xmin / 1000 * w)
    y0 = int(ymin / 1000 * h)
    x1 = int(xmax / 1000 * w)
    y1 = int(ymax / 1000 * h)

    cx = (x0 + x1) // 2
    cy = (y0 + y1) // 2
    half_w = max((x1 - x0) // 2 + margen_h, min_w // 2)
    half_h = max((y1 - y0) // 2 + margen_v, min_h // 2)

    x0 = max(0, cx - half_w)
    x1 = min(w, cx + half_w)
    y0 = max(0, cy - half_h)
    y1 = min(h, cy + half_h)

    recorte = img.crop((x0, y0, x1, y1))

    if recorte.width < 400:
        scale = 400 / recorte.width
        recorte = recorte.resize(
            (int(recorte.width * scale), int(recorte.height * scale)),
            Image.LANCZOS,
        )

    buf = io.BytesIO()
    recorte.save(buf, format="PNG")
    return buf.getvalue()


def extraer_datos(pdf_bytes):
    img_bytes = pdf_primera_pagina(pdf_bytes)
    response = client.models.generate_content(
        model="gemini-2.0-flash",
        contents=[
            types.Part.from_bytes(data=img_bytes, mime_type="image/png"),
            types.Part.from_text(text=PROMPT),
        ],
    )
    texto = response.text.strip()
    if texto.startswith("```"):
        texto = texto.split("```")[1]
        if texto.startswith("json"):
            texto = texto[4:]
    return json.loads(texto.strip()), img_bytes


def incrustar_imagen(ws, img_bytes, col, row, img_width=350):
    """Incrusta una imagen en la celda indicada del worksheet."""
    img_pil = Image.open(io.BytesIO(img_bytes))
    ratio = img_width / img_pil.width
    new_h = int(img_pil.height * ratio)
    img_pil = img_pil.resize((img_width, new_h), Image.LANCZOS)
    img_buf = io.BytesIO()
    img_pil.save(img_buf, format="PNG")
    img_buf.seek(0)

    xl_img = XLImage(img_buf)
    xl_img.width = img_width
    xl_img.height = new_h
    ws.add_image(xl_img, f"{get_column_letter(col)}{row}")


def generar_excel(resultados, recortes_cuenta, recortes_nombre):
    """Genera el Excel en memoria y lo devuelve como bytes."""
    columnas = ["archivo", "nombre_completo", "cuenta", "clabe", "banco", "tipo"]
    df = pd.DataFrame(resultados)[columnas]

    buf_excel = io.BytesIO()
    df.to_excel(buf_excel, index=False)
    buf_excel.seek(0)

    wb = load_workbook(buf_excel)
    ws = wb.active

    COL_NOMBRE = 8
    COL_CUENTA = 9
    IMG_WIDTH = 350
    ROW_HEIGHT = 90

    ws.cell(row=1, column=COL_NOMBRE, value="recorte_nombre")
    ws.cell(row=1, column=COL_CUENTA, value="recorte_cuenta")
    ws.column_dimensions[get_column_letter(COL_NOMBRE)].width = 50
    ws.column_dimensions[get_column_letter(COL_CUENTA)].width = 50

    for idx, fila in enumerate(resultados, start=2):
        ws.row_dimensions[idx].height = ROW_HEIGHT
        nombre_bytes = recortes_nombre.get(fila["archivo"])
        cuenta_bytes = recortes_cuenta.get(fila["archivo"])
        if nombre_bytes:
            incrustar_imagen(ws, nombre_bytes, COL_NOMBRE, idx, IMG_WIDTH)
        if cuenta_bytes:
            incrustar_imagen(ws, cuenta_bytes, COL_CUENTA, idx, IMG_WIDTH)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ── Interfaz ─────────────────────────────────────────────────────────
archivos = st.file_uploader(
    "Sube los PDFs de carátulas bancarias",
    type=["pdf"],
    accept_multiple_files=True,
)

if archivos:
    st.info(f"📂 {len(archivos)} archivo(s) seleccionado(s)")

    if st.button("🚀 Procesar", type="primary"):
        resultados = []
        recortes_cuenta = {}
        recortes_nombre = {}
        barra = st.progress(0, text="Iniciando...")

        for i, archivo in enumerate(archivos):
            nombre = archivo.name
            barra.progress(
                (i) / len(archivos),
                text=f"Procesando {nombre} ({i+1}/{len(archivos)})...",
            )

            try:
                pdf_bytes = archivo.read()
                datos, img_bytes = extraer_datos(pdf_bytes)

                fila = {"archivo": nombre}
                for campo in ["nombre_completo", "cuenta", "clabe", "banco", "tipo"]:
                    v = datos.get(campo, {}).get("valor", "Verificar")
                    fila[campo] = v if v else "Verificar"
                resultados.append(fila)

                bbox_cuenta = datos.get("cuenta", {}).get("bbox", [0, 0, 0, 0])
                recorte_cuenta = recortar_zona(img_bytes, bbox_cuenta)
                if recorte_cuenta:
                    recortes_cuenta[nombre] = recorte_cuenta

                bbox_nombre = datos.get("nombre_completo", {}).get("bbox", [0, 0, 0, 0])
                recorte_nombre = recortar_zona(img_bytes, bbox_nombre)
                if recorte_nombre:
                    recortes_nombre[nombre] = recorte_nombre

                # Mostrar resultado individual
                with st.expander(f"✅ {nombre}", expanded=False):
                    st.markdown(f"**Nombre:** {fila['nombre_completo']}")
                    st.markdown(f"**Cuenta:** {fila['cuenta']}")
                    st.markdown(f"**CLABE:** {fila['clabe']}")
                    st.markdown(f"**Banco:** {fila['banco']}")
                    st.markdown(f"**Tipo:** {fila['tipo']}")
                    c1, c2 = st.columns(2)
                    with c1:
                        if recorte_nombre:
                            st.image(recorte_nombre, caption="Recorte de nombre")
                    with c2:
                        if recorte_cuenta:
                            st.image(recorte_cuenta, caption="Recorte de cuenta")

            except Exception as e:
                st.error(f"❌ Error en {nombre}: {e}")
                resultados.append({
                    "archivo": nombre,
                    "nombre_completo": "Verificar",
                    "cuenta": "Verificar",
                    "clabe": "Verificar",
                    "banco": "Verificar",
                    "tipo": "Verificar",
                })

        barra.progress(1.0, text="✅ Listo")

        if resultados:
            st.success(f"Se procesaron {len(resultados)} archivo(s).")

            # Tabla resumen
            st.dataframe(
                pd.DataFrame(resultados),
                use_container_width=True,
                hide_index=True,
            )

            # Generar y ofrecer descarga del Excel
            excel_bytes = generar_excel(resultados, recortes_cuenta, recortes_nombre)
            st.download_button(
                label="📥 Descargar Excel",
                data=excel_bytes,
                file_name="cuentas_extraidas.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
            )
